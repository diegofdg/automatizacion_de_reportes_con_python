import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

archivo_excel = pd.read_excel('project-report/supermarket_sales.xlsx')
# print(archivo_excel[['Gender', 'Product line', 'Total']])

table_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
print(table_pivote)

table_pivote.to_excel('sales_2022.xlsx', startrow=4, sheet_name='Report')

wb = load_workbook('sales_2022.xlsx')
pestania = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

print('Columna', min_col)
print('Columna', max_col)
print('Fila', min_fila)
print('Fila', max_fila)

### Gráficos
barChart = BarChart()

data = Reference(pestania, min_col = min_col +1, max_col = max_col, min_row = min_fila, max_row = max_fila)
categoria = Reference(pestania, min_col = min_col, max_col = min_col, min_row = min_fila +1, max_row = max_fila)

barChart.add_data(data, titles_from_data=True)
barChart.add_data(categoria)

pestania.add_chart(barChart, 'B12')
barChart.title = 'Ventas'
barChart.style = 4

pestania['A1'] = 'Reporte'
pestania['A2'] = '2022'

pestania['A1'].font = Font('Arial', bold=True, size=20)
pestania['A2'].font = Font('Arial', bold=True, size=15)

wb.save('sales_2022.xlsx')